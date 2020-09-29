import openpyxl as xl

main_wb = xl.load_workbook('bsa.xlsx')  # Download a database

sheet = main_wb['Sheet1']  # Choose a sheet

cell = sheet['b2']  # Choose exact cell (option.1)

cell1 = sheet.cell(3, 2)  # Choose exact cell (option.2)

max_row = sheet.max_row  # Max_row ...

list_1 = []  # list for collecting one exact column
full_list = []  # list for collecting few exact columns
columns = 0
for column in range(1, sheet.max_column):
    cell = sheet.cell(4, column)
    if cell.value is None:
        continue
    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row, column)
        list_1.append(cell.value)   # Get column
    else:
        full_list.append(list_1)    # write it in the full_list
        columns += 1
        list_1 = []     # clear a recorded column


sheet2 = main_wb['Sheet2']  # create new sheet to copy new version of sheet.1

for column in range(1, columns + 1):
    for row in range(2, len(full_list[0]) + 2):
        cell = sheet2.cell(row, column)
        cell.value = full_list[column-1][row-2]


main_wb.save('Added_new_db7.xlsx')  # I got excel-sheet in normal format

main_wb = xl.load_workbook('Added_new_db7.xlsx')  # Download a database

sheet = main_wb['Sheet2']

list_1 = []  # list for collecting one exact column
full_list = []  # list for collecting few exact columns
columns = 0
need_only = [1, 3, 7, 8, 9, 10, 11, 12, 13, 15, 17, 19, 21]  # index of columns that i need
for column in need_only:
    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row, column)
        list_1.append(cell.value)   # Get column
    else:
        full_list.append(list_1)    # write it in the full_list
        columns += 1
        list_1 = []     # clear a recorded column


sheet2 = main_wb['Sheet3']  # create new sheet to copy new version of sheet.2

for column in range(1, columns + 1):
    for row in range(2, len(full_list[0]) + 2):
        cell = sheet2.cell(row, column)
        cell.value = full_list[column-1][row-2]

main_wb.save('Added_new_db8.xlsx')  # I got excel-sheet in normal format

main_wb = xl.load_workbook('Added_new_db8.xlsx')  # Download a database
sheet = main_wb['Sheet3']
Sheet2 = main_wb['Sheet2']
cell = sheet.cell(5, 6)
main_wb.remove(Sheet2)  # Delete Sheet2!!!
#
# main_wb.save('Removed Sheet 2.xlsx')
# listtt = []
# list = [1, 2, 3]
# listt = [4, 5, 6]
# listtt.append(list)
# listtt.append(listt)
#
# print(listtt)
#
# for row in range(2, max_row + 1):  # I want to add a new column
#     cell = sheet.cell(row, 2)  # Choose exact cell
#     corrected_price = cell.value * 1.1  # Correcting a price
#     corrected_price_cell = sheet.cell(row, 3)  # Get index of cell that I need to fill
#     corrected_price_cell.value = corrected_price  # Fill cell
#
# column_name = sheet.cell(1, 3)
# column_name.value = 'After a time...'  # Give new name to fourth column
#
# main_wb.save('new_wb.xlsx')  # Save a new database with new value
