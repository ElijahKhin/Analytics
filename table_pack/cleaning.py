import openpyxl as xl

# Next function will remove all empty columns and rows, but be careful, some of the rows may be displayed incorrectly!


def del_empty_clm(file_name, start_row, sheet):
    wb = xl.load_workbook(f'{file_name}.xlsx')
    sheet = wb[sheet]

    lst = []  # list for collecting one exact column
    full_lst = []  # list for collecting few exact columns
    rows_lst = []  # Each column may have different quantity of filled rows, so I need to track it
    columns = 0
    rows = 0
    for column in range(1, sheet.max_column):
        cell = sheet.cell(start_row, column)
        #  Clean empty columns
        if cell.value is None:
            continue
        for row in range(3, sheet.max_row + 1):
            cell = sheet.cell(row, column)
            #  Clean empty rows
            if cell.value is None:
                continue
            #  If value of cell is not Null the cell goes to list
            cell = sheet.cell(row, column)
            lst.append(cell.value)  # Get column
            rows += 1
        else:
            full_lst.append(lst)  # write it in the full_lst
            columns += 1
            rows_lst.append(rows)
            lst = []  # clear a recorded column
            rows = 0  # clear a recorded rows

    wb.create_sheet('Sheet2', 1)
    sheet2 = wb['Sheet2']  # create new sheet to copy new version of sheet.1

    for column in range(1, columns + 1):  # Recording updated sheet
        for row in range(2, rows_lst[column-1] + 2):
            cell = sheet2.cell(row, column)
            cell.value = full_lst[column-1][row-2]
        if max(rows_lst) != rows_lst[column - 1]:
            cell = sheet2.cell(3, column)
            k = cell.value
            cell.value = str(k) + " " + '(displayed incorrectly)'

    wb.remove(sheet)
    new_file_name = 'clean_table_v5.xlsx'
    wb.save(new_file_name)
    return new_file_name


def choose_clm(file_name, sheet):
    wb = xl.load_workbook(file_name)
    sheet = wb[sheet]

    list_1 = []  # list for collecting one exact column
    full_list = []  # list for collecting few exact columns
    columns = 0
    need_only = [1, 3, 7, 8, 9, 10, 11, 12, 13, 15, 17, 19, 21]  # index of columns that i need
    for column in need_only:
        for row in range(2, sheet.max_row + 1):
            cell = sheet.cell(row, column)
            list_1.append(cell.value)  # Get column
        else:
            full_list.append(list_1)  # write it in the full_list
            columns += 1
            list_1 = []  # clear a recorded column
    wb.create_sheet('Sheet3', 2)
    sheet = wb['Sheet3']
    for column in range(1, columns + 1):
        for row in range(2, len(full_list[0]) + 2):
            cell = sheet.cell(row, column)
            cell.value = full_list[column - 1][row - 2]
    file_name = 'exact_clm.xlsx'
    sheet = wb['Sheet2']
    wb.remove(sheet)
    wb.save(file_name)
    return file_name


def change_exact_column(file_name, number_of_column, sheet):
    wb = xl.load_workbook(file_name)
    sheet = wb[sheet]
    n = 1
    for row in range(4, sheet.max_row + 1):
        cell = sheet.cell(row, number_of_column)
        cell.value = n
        n += 1
    wb.save(file_name)
    return print(file_name)



